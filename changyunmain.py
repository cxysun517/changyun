# ##################################
# python操作excel的4个工具包：
# xlrd: 对.xls进行读相关操作
# xlwt: 对.xls进行写相关操作
# xlutils: 对.xls读写操作的整合
# openpyxl：对.xlsx进行读写操作
# 前三个库都只能操作.xls不能操作.xlsx
# 最后一个只能操作.xlsx不能操作.xls
# ##################################

import os.path  # 文件及目录操作模块
import shutil  # python3.8的内置模块，文件的复制、移动、删除、压缩、解压等操作
import sys  # 主要负责与Python解释器进行交互,该模块提供了一系列用于控制Python运行环境的函数和变量

from openpyxl import load_workbook  # load_workbook()函数，打开已有工作簿即一个excel文件，不能创建
from openpyxl.styles import Border, Side, PatternFill, Alignment

# 单元格设置
# 边框：Border, Side
# 单元格填充：PatternFill图案填充、GradientFill渐变色填充
# 字体：Font
# 对齐方式：Alignment

# 家庭ID
family1 = 2661243  # 嵌入式测试-Zigbee-地下B1穿墙长运
family2 = 11651573  # 嵌入式测试-BLE-7楼屏蔽室长运
family3 = 22073540  # 嵌入式测试-zigbee-7E小网关家庭长运
family4 = 18367531  # 嵌入式测试-BLE-7E单插网关长运
family5 = 69842907  # 嵌入式测试-7E-ssd212双联蓝牙网关家庭长运

# 初始化
row1 = 5  # family1初始从第5行写入具体的告警信息数据，第4行为告警信息标题，第2和3行合并为家庭标题
row2 = 10  # family2初始从第10行写入具体的告警信息数据，第9行为告警信息标题，第7和8行合并为家庭标题
row3 = 15
row4 = 20
row5 = 25

with open('test.txt', 'r', encoding='utf-8') as f:
    read_data = f.read()
    a = read_data.split('【长运设备异常报警】')  # 指定分隔符对字符串进行切片，并返回分割后的字符串列表
    print(a)
    # 告警信息字符串分割后返回的列表的长度
    # 注：若没有告警产生，test.txt文件内容为空，则返回有一个空字符串元素的列表['']，长度为1
    length = len(a)

filename = '网关长运环境测试数据.xlsx'
wb = load_workbook(filename)  # 打开excel
sheet = wb.create_sheet('Mysheet')  # 创建名称为Mysheet的工作表
ws = wb['Mysheet']  # 获取名称为Mysheet的工作表对象，工作表中行、列的index都是从1开始

# 打印5个长运环境离线数据的标题并设置格式
flag1 = 0
flag2 = 0
flag3 = 0
flag4 = 0
flag5 = 0

# 单元格边框设置
border_set = Border(left=Side(style='medium', color='000000'),
                    right=Side(style='medium', color='000000'),
                    top=Side(style='medium', color='000000'),
                    bottom=Side(style='medium', color='000000'))

# 长运环境离线数据的标题单元格填充
# PatternFill参数：
# patternType：填充图案类型，纯色填充为solid，为none时无填充
# fgColor：图案前景色
# bgColor：图案背景色
fill1 = PatternFill("solid", fgColor="AACF91")  # 绿色
fill2 = PatternFill("solid", fgColor="1874CD")  # 蓝色

# 告警信息的标题
t = ['虚拟ID', '设备名称', '上报类型', '原因', '日期', '时间', '后台日志', '统计数据']

i = 0
# 针对只有一次告警的情况，最后1次告警时间赋值为空
end_date = ''

# 新建工作表按家庭分类添加告警信息，一条信息一行，第1条信息的前面1行填写告警信息标题 #
while True:
    # 获取第一次告警的日期
    if i == 0:
        start_date = a[i].split()  # split()不指定参数即默认分隔符为空格、换行符\n、制表符\t，分割次数不限
        i = 1

    # 处理每次告警信息
    # 若没有告警产生，length=1，则不进入此分支
    # 若有告警产生（只有1条告警时，length=2），则进入此分支
    if i < length:

        b = a[i].split()
        while b[10] != '上报类型:':  # 校验设备名称是否有空格,导致list错位

            b[9] = b[9] + b[10]
            del (b[10])

        # 获取家庭ID
        fid = int(b[7])

        # 如果告警信息来自family1
        if fid == family1:
            if flag1 == 0:
                for j in range(2, 10):
                    # 单元格操作
                    ws.cell(row=row1 - 1, column=j,
                            value=t[j - 2])  # 第4行，从第2列到第9列，依次填入虚拟ID,设备名称,上报类型,原因,日期,时间,后台日志,统计数据
                    ws.cell(row=row1 - 1, column=j).border = border_set  # 边框设置
                    ws.cell(row1 - 1, j).fill = fill2  # 填充蓝色
                flag1 += 1

            # 初始从第5行开始填写具体告警信息
            ws.cell(row=row1, column=2, value=b[3])
            ws.cell(row=row1, column=3, value=b[9])
            ws.cell(row=row1, column=4, value=b[11])
            ws.cell(row=row1, column=5, value=b[13])
            ws.cell(row=row1, column=6, value=b[15])
            ws.cell(row=row1, column=7, value=b[16])
            ws.cell(row=row1, column=8, value=b[18])
            ws.cell(row=row1, column=9, value=b[20])

            for j in range(2, 10):
                ws.cell(row=row1, column=j).border = border_set

            # 给原第row1+1行前面插入1行
            # 工作表每写入1条家庭1的告警信息，在该行告警信息后插入新的一行
            # 始终保证工作表每个家庭最后1条告警信息与下一个家庭标题行间隔2行
            ws.insert_rows(row1 + 1, 1)

            # 工作表每写入1条家庭1的告警信息，计算下一条告警信息写入的行数
            row1 += 1  # 家庭1当前写入的最后1条告警信息所在行的下一行
            row2 += 1  # 家庭2当前写入的最后1条告警信息所在行的下一行
            row3 += 1  # 家庭3当前写入的最后1条告警信息所在行的下一行
            row4 += 1  # 家庭4当前写入的最后1条告警信息所在行的下一行
            row5 += 1  # 家庭5当前写入的最后1条告警信息所在行的下一行

        if fid == family2:
            if flag2 == 0:
                for j in range(2, 10):
                    ws.cell(row=row2 - 1, column=j, value=t[j - 2])
                    ws.cell(row=row2 - 1, column=j).border = border_set
                    ws.cell(row2 - 1, j).fill = fill2
                flag2 += 1

            # 初始从第10行开始填写具体告警信息
            ws.cell(row=row2, column=2, value=b[3])
            ws.cell(row=row2, column=3, value=b[9])
            ws.cell(row=row2, column=4, value=b[11])
            ws.cell(row=row2, column=5, value=b[13])
            ws.cell(row=row2, column=6, value=b[15])
            ws.cell(row=row2, column=7, value=b[16])
            ws.cell(row=row2, column=8, value=b[18])
            ws.cell(row=row2, column=9, value=b[20])

            for j in range(2, 10):
                ws.cell(row=row2, column=j).border = border_set

            ws.insert_rows(row2 + 1)

            row2 += 1
            row3 += 1
            row4 += 1
            row5 += 1

        if fid == family3:
            if flag3 == 0:
                for j in range(2, 10):
                    ws.cell(row=row3 - 1, column=j, value=t[j - 2])
                    ws.cell(row=row3 - 1, column=j).border = border_set
                    ws.cell(row3 - 1, j).fill = fill2
                flag3 += 1

            # 初始从第15行开始填写具体告警信息
            ws.cell(row=row3, column=2, value=b[3])
            ws.cell(row=row3, column=3, value=b[9])
            ws.cell(row=row3, column=4, value=b[11])
            ws.cell(row=row3, column=5, value=b[13])
            ws.cell(row=row3, column=6, value=b[15])
            ws.cell(row=row3, column=7, value=b[16])
            ws.cell(row=row3, column=8, value=b[18])
            ws.cell(row=row3, column=9, value=b[20])

            for j in range(2, 10):
                ws.cell(row=row3, column=j).border = border_set

            ws.insert_rows(row3 + 1)

            row3 += 1
            row4 += 1
            row5 += 1

        if fid == family4:
            if flag4 == 0:
                for j in range(2, 10):
                    ws.cell(row=row4 - 1, column=j, value=t[j - 2])
                    ws.cell(row=row4 - 1, column=j).border = border_set
                    ws.cell(row4 - 1, j).fill = fill2
                flag4 += 1

            ws.cell(row=row4, column=2, value=b[3])
            ws.cell(row=row4, column=3, value=b[9])
            ws.cell(row=row4, column=4, value=b[11])
            ws.cell(row=row4, column=5, value=b[13])
            ws.cell(row=row4, column=6, value=b[15])
            ws.cell(row=row4, column=7, value=b[16])
            ws.cell(row=row4, column=8, value=b[18])
            ws.cell(row=row4, column=9, value=b[20])

            for j in range(2, 10):
                ws.cell(row=row4, column=j).border = border_set

            ws.insert_rows(row4 + 1)

            row4 += 1
            row5 += 1

        if fid == family5:
            if flag5 == 0:
                for j in range(2, 10):
                    ws.cell(row=row5 - 1, column=j, value=t[j - 2])
                    ws.cell(row=row5 - 1, column=j).border = border_set
                    ws.cell(row5 - 1, j).fill = fill2
                flag5 += 1

            ws.cell(row=row5, column=2, value=b[3])
            ws.cell(row=row5, column=3, value=b[9])
            ws.cell(row=row5, column=4, value=b[11])
            ws.cell(row=row5, column=5, value=b[13])
            ws.cell(row=row5, column=6, value=b[15])
            ws.cell(row=row5, column=7, value=b[16])
            ws.cell(row=row5, column=8, value=b[18])
            ws.cell(row=row5, column=9, value=b[20])

            for j in range(2, 10):
                ws.cell(row=row5, column=j).border = border_set

            ws.insert_rows(row5 + 1)

            row5 += 1

        # 得到最后一个告警的日期
        if i + 2 == length:
            date = a[i].split('数据平台')
            end_date = date[1].split()

        # 每循环处理告警信息1次后，i递增，直到处理完所有告警信息，break退出while循环
        i += 1

    else:
        break
# 新建工作表按家庭分类添加告警信息，一条信息一行，第1条信息的前面1行填写告警信息标题 #

# sheet名称及execl文件名称设置 #
# 若没有产生告警信息，退出程序
if len(start_date) == 0:
    print("无告警信息产生，中止程序运行")
    sys.exit()

# 若只有1条告警信息，sheet名称及execl文件名称设置如下
if len(start_date) != 0 and len(end_date) == 0:
    ws.title = start_date[1] + '离线数据'
    f_name_date = start_date[1]

# 若大于1条告警信息，sheet名称及execl文件名称设置如下
if len(start_date) != 0 and len(end_date) != 0:
    ws.title = start_date[1] + '至' + end_date[0] + '离线数据'
    f_name_date = start_date[1] + '至' + end_date[0]
# sheet名称及execl文件名称设置 #

# 删除没告警的家庭占用行 #
colA = ws['B']  # 获取工作表Mysheet的B列对象
k = 0
index = ['', '', '', '', '']
# 按家庭统计告警信息标题：虚拟ID,设备名称,上报类型,原因,日期,时间,后台日志,统计数据 所在的行数
for cell in colA:
    if cell.value == '虚拟ID':
        index[k] = cell.row
        k += 1
# 告警家庭数<=5，列表x存储产生告警的家庭其告警信息标题所在行数
x = list(index[:k])

# flag1 == 0 ，则家庭1没告警，删除给家庭1预留的行
if flag1 == 0:
    if len(x) != 0:  # 其它家庭有告警
        ws.delete_rows(x[0] - 7, 5)
        # 1st参数代表从哪行开始，2nd代表删除几行
        # 例如：家庭1和2都没告警，家庭3有告警，其告警信息标题在第14行，即x[0]=14，12和13两行预留合并填写家庭名称
        # 则删除工作表中第7行至第11行（原本用于家庭2），家庭3上移至从第7行开始
    # 每当删除一个家庭时，x列表要更新，元素相应的都减去5
    for m in range(len(x)):
        x[m] = x[m] - 5
    # 此时x[0]=9
else:
    del (x[0])

# flag2 == 0 ，则家庭2没告警，删除给家庭2预留的行
if flag2 == 0:
    if len(x) != 0:
        ws.delete_rows(x[0] - 7, 5)
        # 删除工作表中第2行至第6行（原本用于家庭1），家庭3再次上移至从第2行开始
    for m in range(len(x)):
        x[m] = x[m] - 5
else:
    del (x[0])

if flag3 == 0:
    if len(x) != 0:
        ws.delete_rows(x[0] - 7, 5)
    for m in range(len(x)):
        x[m] = x[m] - 5
else:
    del (x[0])

if flag4 == 0:
    if len(x) != 0:
        ws.delete_rows(x[0] - 7, 5)
    for m in range(len(x)):
        x[m] = x[m] - 5
else:
    del (x[0])

if flag5 == 0:
    if len(x) != 0:
        ws.delete_rows(x[0] - 7, 5)
    for m in range(len(x)):
        x[m] = x[m] - 5
else:
    del (x[0])

# 删除没告警的家庭占用行 #

# 合并单元格设置家庭标题并填充绿色 #
list = ['嵌入式测试-Zigbee-地下B1穿墙长运', '嵌入式测试-BLE-7楼屏蔽室长运', '嵌入式测试-zigbee-7E小网关家庭长运', '嵌入式测试-BLE-7E单插网关长运',
        '嵌入式测试-7E-ssd212双联蓝牙网关长运']
for cell in colA:
    if cell.value == '虚拟ID':
        # 合并单元格，行范围：虚拟ID所在行的前面2行 列范围：2至9列，即B至I列
        ws.merge_cells(start_row=cell.row - 2, start_column=2, end_row=cell.row - 1, end_column=9)
        # 判断属于哪个家庭，合并单元格并赋值
        if flag1 == 1:
            # 合并单元格赋值：嵌入式测试-Zigbee-地下B1穿墙长运
            # 单元格坐标：行数cell.row-2，列数2即B列
            ws.cell(row=cell.row - 2, column=2, value=list[0])
            flag1 += 1
        else:
            if flag2 == 1:
                ws.cell(row=cell.row - 2, column=2, value=list[1])
                flag2 += 1
            else:
                if flag3 == 1:
                    ws.cell(row=cell.row - 2, column=2, value=list[2])
                    flag3 += 1
                else:
                    if flag4 == 1:
                        ws.cell(row=cell.row - 2, column=2, value=list[3])
                        flag4 += 1
                    else:
                        if flag5 == 1:
                            ws.cell(row=cell.row - 2, column=2, value=list[4])
                            flag5 += 1

        # 设置合并单元格的边框和填充色
        r = cell.row - 2
        while r < cell.row:
            for c in range(2, 10):
                ws.cell(r, c).border = border_set
                ws.cell(r, c).fill = fill1
            r += 1

# 合并单元格设置家庭标题并填充绿色 #

# 设置所有单元格横向纵向居中 #
max_r = ws.max_row  # 获取最大行
min_r = ws.min_row  # 获取最小行
max_c = ws.max_column  # 获取最大列
min_c = ws.min_column  # 获取最小列
# 方法一：
# for colA in [ws['B'],ws['C'],ws['D'],ws['E'],ws['F'],ws['G'],ws['H'],ws['I']]:
#     for cell in colA:
#         cell.alignment = Alignment(horizontal='center', vertical='center')

# 方法二：
for row in ws.iter_rows(min_row=min_r, max_row=max_r):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 设置所有单元格横向纵向居中 #

# 设置列宽 #
ws.column_dimensions['B'].width = 25.0  # 列宽
ws.column_dimensions['C'].width = 30.0
ws.column_dimensions['F'].width = 15.0
ws.column_dimensions['G'].width = 10.0
ws.column_dimensions['H'].width = 90.0
ws.column_dimensions['I'].width = 60.0
# 设置列宽 #
wb.save(filename)  # 保存变更
wb.close()

# 路径拼接
dst_file = os.path.join(r'C:\长运环境', '网关长运环境测试数据' + f_name_date + '.xlsx')
print(dst_file)
src_file = os.path.join(sys.path[0], '网关长运环境测试数据.xlsx')  # sys.path[0]获取脚本运行所在目录
print(src_file)
# 1st参数：需要复制的源文件的文件路径+文件名    2st参数：目标文件的文件路径+文件名
shutil.copyfile(src_file, dst_file)
